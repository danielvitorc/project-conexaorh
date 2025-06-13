from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image
from ..models import RequisicaoPessoal
from django.contrib.auth import get_user_model

User = get_user_model()


def preencher_excel_rp(rp):
    wb = load_workbook("conexaorh/static/rp_modelo_atualizado.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=60):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    template_processo_seletivo = "[  ] INTERNO  [  ] MISTO  [  ] EXTERNO  [  ] SIGILOSO"

    processo_seletivo = [(v or "").strip().casefold() for v in (rp.processo_seletivo or "").split(",")]

    for opcao_processo_seletivo in ["INTERNO", "MISTO", "EXTERNO", "SIGILOSO"]:
        if opcao_processo_seletivo.casefold() in processo_seletivo:
            template_processo_seletivo = template_processo_seletivo.replace(f"[  ] {opcao_processo_seletivo}", f"[X] {opcao_processo_seletivo}")

    # Texto base da célula no Excel
    template_sexo = "[  ] Masculino   [  ] Feminino   [  ] Indiferente"

    # Pega os valores separados por vírgula, normaliza com strip() e title()
    sexos = [(s or "").strip().title() for s in (rp.sexo or "").split(",")]

    # Para cada opção, substitui se estiver presente na lista
    for opcao in ["Masculino", "Feminino", "Indiferente"]:
        if opcao in sexos:
            template_sexo = template_sexo.replace(f"[  ] {opcao}", f"[X] {opcao}")
    
    template_beneficios = "[  ] Plano Unimed Cnu      [  ] Plano Saúde Hapvida      [  ] Alelo Mobilidade     [  ] Seguro De Vida"

    beneficios = [(b or "").strip().title() for b in (rp.beneficios or "").split(",")]

    for opcao_beneficios in ["Plano Unimed Cnu", "Plano Saúde Hapvida", "Alelo Mobilidade", "Seguro De Vida"]:
        for opcao_beneficios in beneficios:
            template_beneficios = template_beneficios.replace(f"[  ] {opcao_beneficios}", f"[X] {opcao_beneficios}") 
    
    template_viagem = "[  ] Sim [  ] Não"

    viagens = [(v or "").strip().casefold() for v in (rp.exige_viagem or "").split(",")]

    for opcao in ["Sim", "Não"]:
        if opcao.casefold() in viagens:
            template_viagem = template_viagem.replace(f"[  ] {opcao}", f"[X] {opcao}")

    template_cnh = "[  ] Sim [  ] Não"

    cnh = [(v or "").strip().casefold() for v in (rp.cnh or "").split(",")]

    for opcao_cnh in ["Sim", "Não"]:
        if opcao_cnh.casefold() in cnh:
            template_cnh = template_cnh.replace(f"[  ] {opcao_cnh}", f"[X] {opcao_cnh}")

    ws["I4"] = rp.data_solicitacao.strftime("%d/%m/%Y") if rp.data_solicitacao else ""
    ws["G7"] = rp.requisitante or ""
    ws["X7"] = rp.cargo or ""
    ws["G9"] = rp.salario or ""
    ws["U9"] = rp.adicionais or ""
    ws["AF9"] = rp.quantidade_vagas or ""
    ws["G11"] = rp.horario_trabalho_inicio.strftime("%H:%M") if rp.horario_trabalho_inicio else ""
    ws["P11"] = rp.tipo_ponto or ""
    ws["X11"] = rp.inicio_contrato.strftime("%d/%m/%Y") if rp.inicio_contrato else ""
    ws["AF11"] = rp.termino_contrato.strftime("%d/%m/%Y") if rp.termino_contrato else ""
    ws["G13"] = rp.tipo_contratacao or ""
    ws["G15"] = template_beneficios
    ws["U13"] = rp.motivo_contracao or ""
    ws["H17"] = rp.subtituicao or ""
    ws["AE17"] = rp.matricula or "" 
    ws["I19"] = rp.justificativa_substituicao or ""
    ws["AC19"] = rp.justificativa_outros or ""
    ws["E22"] = template_processo_seletivo  
    #ws["E22"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws["Y22"] = rp.localidade or ""
    ws["AG22"] = rp.base or ""
    ws["O24"] = template_sexo
    ws["AD24"] = template_viagem
    ws["C26"] = template_cnh
    ws["K26"] = rp.tipo_cnh or ""
    ws["S26"] = rp.outros_cnh or ""
    ws["G28"] = rp.departamento or ""
    ws["Y28"] = rp.escolaridade or ""
    ws["G30"] = rp.gestor_imediato or ""
    ws["Y30"] = rp.centro_custo or ""
    ws["G32"] = rp.cursos or ""
    ws["G34"] = rp.experiencias or ""
    ws["A38"] = rp.habilidades_comportamentais or ""
    ws["S38"] = rp.principais_atribuicoes or ""
    ws["G53"] = rp.candidato_aprovado or ""
    ws["AB53"] = rp.n_rp or ""
    ws["A58"] = rp.data_autorizacao_rh.strftime("%d/%m/%Y") if rp.data_autorizacao_rh else ""
    ws["S58"] = rp.data_autorizacao_diretor.strftime("%d/%m/%Y") if rp.data_autorizacao_diretor else ""
    ws["AB58"] = rp.data_autorizacao_presidente.strftime("%d/%m/%Y") if rp.data_autorizacao_presidente else ""
    ws["J58"] = rp.data_autorizacao_gestor.strftime("%d/%m/%Y") if rp.data_autorizacao_gestor else ""

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, rp.assinatura_diretor, "S61")
    adicionar_imagem_excel(ws, rp.assinatura_presidente, "AB61")
    adicionar_imagem_excel(ws, rp.assinatura_rh, "A61")
    adicionar_imagem_excel(ws, rp.assinatura_gestor, "J61")

    # Salva em memória
    buffer = io.BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_rp_excel(request, registro_id):
    registro = get_object_or_404(RequisicaoPessoal, id=registro_id)

    excel_bytes = preencher_excel_rp(registro)
    nome_arquivo = f"RP_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response