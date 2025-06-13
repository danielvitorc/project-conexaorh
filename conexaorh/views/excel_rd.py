from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image
from ..models import RequisicaoDesligamento

     
def preencher_excel_rd(rd):
    wb = load_workbook("conexaorh/static/rd_modelo.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=60):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    template_tipo_desligamento = (
        "[  ] SEM JUSTA CAUSA   [  ] POR JUSTA CAUSA   [  ] COMUM ACORDO\n"
        "[  ] TÉRMINO DE CONTRATO   [  ] RESCISÃO ANTECIPADA CONTRATO DE EXPERIÊNCIA")
              
    tipo_desligamento = [(v or "").strip().casefold() for v in (rd.tipo_desligamento or "").split(",")]

    for opcao_tipo_desligamento in ["SEM JUSTA CAUSA", "POR JUSTA CAUSA", "COMUM ACORDO", "TÉRMINO DE CONTRATO", 
        "RESCISÃO ANTECIPADA CONTRATO DE EXPERIÊNCIA"]:
        if opcao_tipo_desligamento.casefold() in tipo_desligamento:
            template_tipo_desligamento = template_tipo_desligamento.replace(f"[  ] {opcao_tipo_desligamento}", f"[X] {opcao_tipo_desligamento}")

    template_motivo_desligamento = (
        "[  ] À PEDIDO DO COLABORADOR               [  ] REDUÇÃO DE QUADRO               [  ] PROBLEMAS COM SUPERIORES\n"
        "[  ] INDISCIPLINA               [  ] EXCESSO DE FALTAS               [  ] INADEQUAÇÃO DO PROFISSIONAL À FUNÇÃO\n"
        "[  ] RELACIONAMENTO COM A EQUIPE INADEQUADO               [  ] BAIXO DESEMPENHO               [  ] DESMOBILIZAÇÃO\n"
        "[  ] REDUÇÃO DE CUSTO")
              
    motivo_desligamento = [(v or "").strip().casefold() for v in (rd.motivo_desligamento or "").split(",")]

    for opcao_motivo_desligamento in ["À PEDIDO DO COLABORADOR", "REDUÇÃO DE QUADRO", "PROBLEMAS COM SUPERIORES", 
        "INDISCIPLINA", "EXCESSO DE FALTAS", "INADEQUAÇÃO DO PROFISSIONAL À FUNÇÃO", "RELACIONAMENTO COM A EQUIPE INADEQUADO", 
        "BAIXO DESEMPENHO", "DESMOBILIZAÇÃO", "REDUÇÃO DE CUSTO"]:
        if opcao_motivo_desligamento.casefold() in motivo_desligamento:
            template_motivo_desligamento = template_motivo_desligamento.replace(f"[  ] {opcao_motivo_desligamento}", f"[X] {opcao_motivo_desligamento}")

    template_tipo_aviso = (
        "[  ] AVISO PRÉVIO TRABALHADO   [  ] AVISO PRÉVIO INDENIZADO")
              
    tipo_aviso = [(v or "").strip().casefold() for v in (rd.tipo_aviso or "").split(",")]

    for opcao_tipo_aviso in ["AVISO PRÉVIO TRABALHADO", "AVISO PRÉVIO INDENIZADO"]:
        if opcao_tipo_aviso.casefold() in tipo_aviso:
            template_tipo_aviso = template_tipo_aviso.replace(f"[  ] {opcao_tipo_aviso}", f"[X] {opcao_tipo_aviso}")

    template_substituicao= (
        "[  ] SIM   [  ] NÃO")
              
    substituicao = [(v or "").strip().casefold() for v in (rd.substituicao or "").split(",")]

    for opcao_substituicao in ["SIM", "NÃO"]:
        if opcao_substituicao.casefold() in substituicao:
            template_substituicao = template_substituicao.replace(f"[  ] {opcao_substituicao}", f"[X] {opcao_substituicao}")

    ws["S4"] = rd.data_solicitacao.strftime("%d/%m/%Y") if rd.data_solicitacao else ""
    ws["G7"] = rd.requisitante or ""
    ws["G9"] = rd.colaborador_desligado or ""
    ws["A12"] = rd.data_desligamento.strftime("%d/%m/%Y") if rd.data_desligamento else ""
    ws["J12"] = rd.funcao or ""
    ws["S12"] = rd.salario or ""
    ws["AB12"] = rd.localidade or ""
    ws["F14"] = rd.matricula or ""
    ws["S14"] = rd.data_admissao.strftime("%d/%m/%Y") if rd.data_admissao else ""
    ws["AD14"] = rd.centro_custo or ""
    ws["A17"] = template_tipo_desligamento
    ws["A17"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A22"] = template_motivo_desligamento
    ws["A22"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A30"] = "[X] OUTRO (CITAR)" if rd.outro_motivo else "[  ] OUTRO (CITAR)"
    ws["A30"].font = Font(bold=True, size=10)
    ws["H30"] = rd.outro_motivo or ""
    ws["H32"] = rd.justificativa_desligamento or ""
    ws["A38"] = template_tipo_aviso
    ws["M46"] = template_substituicao
    ws["H40"] = rd.justificativa_aviso or ""
    ws["V46"] = "[X] BR" if rd.bloqueio_readmissao else "[  ] BR"
    ws["K52"] = rd.data_autorizacao_diretor.strftime("%d/%m/%Y") if rd.data_autorizacao_diretor else ""
    ws["T52"] = rd.data_autorizacao_presidente.strftime("%d/%m/%Y") if rd.data_autorizacao_presidente else ""
    ws["AC52"] = rd.data_autorizacao_rh.strftime("%d/%m/%Y") if rd.data_autorizacao_rh else ""
    ws["B52"] = rd.data_autorizacao_gestor.strftime("%d/%m/%Y") if rd.data_autorizacao_gestor else ""
    

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, rd.assinatura_diretor, "K55")
    adicionar_imagem_excel(ws, rd.assinatura_presidente, "T55")
    adicionar_imagem_excel(ws, rd.assinatura_rh, "AC55")
    adicionar_imagem_excel(ws, rd.assinatura_gestor, "B55")

    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_rd_excel(request, registro_id):
    registro = get_object_or_404(RequisicaoDesligamento, id=registro_id)

    excel_bytes = preencher_excel_rd(registro)
    nome_arquivo = f"Desligamento_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    return response 
     
       
     
     
     

     
     
     
     
     
     
     
      
     
     
     
     
     
     
    