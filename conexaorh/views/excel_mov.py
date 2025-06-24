from django.contrib.auth.decorators import login_required
from django.shortcuts import  get_object_or_404
from django.http import HttpResponse
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image
from ..models import MovimentacaoPessoal
from django.contrib.auth import get_user_model

User = get_user_model()


def preencher_excel_mov(mov):
    wb = load_workbook("conexaorh/static/mov_modelo.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=60):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    try:
        usuario = User.objects.get(username=mov.gestor_proposto)  # Se armazenar username
        nome_gestor = usuario.get_full_name() or usuario.first_name or usuario.username
    except User.DoesNotExist:
        nome_gestor = mov.gestor_proposto  # Usa o texto mesmo se não encontrar

    template_tipo_movimentacao = (
        "[  ] PROMOÇÃO HORIZONTAL   [  ] PROMOÇÃO VERTICAL   [  ] TRANSFERÊNCIA DE ÁREA\n"
        "[  ] TRANSFERÊNCIA DE LOCALIDADE   [  ] TRANSFERÊNCIA DE CONTRATO   [  ] ENQUADRAMENTO")

    tipo_movimentacao = [(v or "").strip().casefold() for v in (mov.tipo_movimentacao or "").split(",")]

    for opcao_tipo_movimentacao in ["PROMOÇÃO HORIZONTAL", "PROMOÇÃO VERTICAL", "TRANSFERÊNCIA DE ÁREA", "TRANSFERÊNCIA DE LOCALIDADE", "TRANSFERÊNCIA DE CONTRATO", "ENQUADRAMENTO"]:
        if opcao_tipo_movimentacao.casefold() in tipo_movimentacao:
            template_tipo_movimentacao = template_tipo_movimentacao.replace(f"[  ] {opcao_tipo_movimentacao}", f"[X] {opcao_tipo_movimentacao}")

    template_tipo_adicional = (
        "[  ] PERICULOSIDADE [  ] INSALUBRIDADE [  ] AJUDA DE CUSTO [  ] ADICIONAL"
    )

    tipo_adicional = [(v or "").strip().casefold() for v in (mov.tipo_adicional or "").split(",")]

    for opcao_tipo_adicional in ["PERICULOSIDADE", "INSALUBRIDADE", "AJUDA DE CUSTO", "ADICIONAL"]:
        if opcao_tipo_adicional.casefold() in tipo_adicional:
            template_tipo_adicional = template_tipo_adicional.replace(f"[  ] {opcao_tipo_adicional}", f"[X] {opcao_tipo_adicional}")

    template_jutificativa_movimentacao = (
        "[  ] REESTRUTURAÇÃO DO DEPARTAMENTO (EMPRESA/UNIDADE) [  ] OPORTUNIDADE DE ASCENÇÃO\n"
        "[  ] INADEQUAÇÃO À ATIVIDADE DO DEPARTAMENTO "
    )

    jutificativa_movimentacao = [(v or "").strip().casefold() for v in (mov.jutificativa_movimentacao or "").split(",")]

    for opcao_jutificativa_movimentacao in ["REESTRUTURAÇÃO DO DEPARTAMENTO (EMPRESA/UNIDADE)", "INADEQUAÇÃO À ATIVIDADE DO DEPARTAMENTO", "OPORTUNIDADE DE ASCENÇÃO"]:
        if opcao_jutificativa_movimentacao.casefold() in jutificativa_movimentacao:
            template_jutificativa_movimentacao = template_jutificativa_movimentacao.replace(f"[  ] {opcao_jutificativa_movimentacao}", f"[X] {opcao_jutificativa_movimentacao}")

    template_substituicao = (
        "[  ] SIM [  ] NÃO  "
    )

    substituicao = [(v or "").strip().casefold() for v in (mov.substituicao or "").split(",")]

    for opcao_substituicao in ["SIM", "NÃO"]:
        if opcao_substituicao.casefold() in substituicao:
            template_substituicao = template_substituicao.replace(f"[  ] {opcao_substituicao}", f"[X] {opcao_substituicao}")
      

    ws["C5"] = mov.n_mov or ""
    ws["O5"] = mov.data_solicitacao.strftime("%d/%m/%Y") if mov.data_solicitacao else ""
    ws["AC5"] = mov.unidade or ""
    ws["A9"] = template_tipo_movimentacao
    ws["A9"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A15"] = "[X] OUTRO (CITAR)" if mov.outro_tipo else "[  ] OUTRO (CITAR)"
    ws["A15"].font = Font(bold=True, size=10)
    ws["H15"] = mov.outro_tipo or ""
    ws["H17"] = mov.colaborador_movimentado or ""
    ws["F19"] = mov.matricula or ""
    ws["S19"] = mov.data_admissao.strftime("%d/%m/%Y") if mov.data_admissao else ""
    ws["AD19"] = mov.outro_info or ""
    ws["A22"] = mov.localidade_atual or ""
    ws["J22"] = mov.cargo_atual or ""
    ws["S22"] = mov.departamento_atual or ""
    ws["AB22"] = mov.salario_atual or ""
    ws["F24"] = mov.gestor_atual or ""
    ws["AD24"] = mov.centro_custo_atual or ""
    ws["A27"] = mov.localidade_proposta or ""
    ws["J27"] = mov.cargo_proposto or ""
    ws["S27"] = mov.departamento_proposto or ""
    ws["AB27"] = mov.salario_proposto or ""
    ws["F29"] = nome_gestor
    ws["AD29"] = mov.centro_custo_proposto or ""
    ws["S32"] = mov.data_movimentacao.strftime("%d/%m/%Y") if mov.data_movimentacao else ""
    ws["A34"] = template_tipo_adicional
    ws["F36"] = mov.tipo_ajuda_custo or ""
    ws["R36"] = mov.valor_ajuda or ""
    ws["AA36"] = mov.periodo or ""
    ws["A39"] = template_jutificativa_movimentacao
    ws["A39"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["H43"] = mov.outro_justificativa or ""
    ws["A43"] = "[X] OUTRO (CITAR)" if mov.outro_justificativa else "[  ] OUTRO (CITAR)"
    ws["A43"].font = Font(bold=True, size=10)
    ws["M45"] = template_substituicao
    ws["H46"] = mov.comentarios or ""
    ws["A32"] = mov.data_solicitacao.strftime("%d/%m/%Y") if mov.data_solicitacao else ""
    ws["A60"] = mov.data_autorizacao_complice.strftime("%d/%m/%Y") if mov.data_autorizacao_complice else ""
    ws["J65"] = mov.data_autorizacao_gestor_proposto.strftime("%d/%m/%Y") if mov.data_autorizacao_gestor_proposto else ""
    ws["S65"] = mov.data_autorizacao_diretor.strftime("%d/%m/%Y") if mov.data_autorizacao_diretor else ""
    ws["AB65"] = mov.data_autorizacao_presidente.strftime("%d/%m/%Y") if mov.data_autorizacao_presidente else ""
    ws["S60"] = mov.data_autorizacao_rh.strftime("%d/%m/%Y") if mov.data_autorizacao_rh else ""
    ws["A65"] = mov.data_autorizacao_gestor_atual.strftime("%d/%m/%Y") if mov.data_autorizacao_gestor_atual else ""
    

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, mov.imagem_assinatura_complice, "A56")
    adicionar_imagem_excel(ws, mov.imagem_assinatura_gestor_proposto, "J68")
    adicionar_imagem_excel(ws, mov.imagem_assinatura_diretor, "S68")
    adicionar_imagem_excel(ws, mov.imagem_assinatura_presidente, "AB68")
    adicionar_imagem_excel(ws, mov.imagem_assinatura_rh, "S56")
    adicionar_imagem_excel(ws,mov.imagem_assinatura_gestor_atual, "A68")


    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_mov_excel(request, registro_id):
    registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)

    excel_bytes = preencher_excel_mov(registro)
    nome_arquivo = f"Movimentação_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response