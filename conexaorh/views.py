from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout
from django.shortcuts import redirect, render, get_object_or_404
from django.utils.timezone import now
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseBadRequest
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import get_user_model
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image
from .forms import RequisicaoPessoalForm, DiretorForm, PresidenteForm, RHForm, MovimentacaoPessoalForm, RequisicaoDesligamentoForm, GestorPropostoApprovalForm, CompliceApprovalForm
from .models import RequisicaoPessoal, MovimentacaoPessoal, RequisicaoDesligamento

User = get_user_model()

def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            # Redirecionamento de acordo com o tipo de usuário
            if user.user_type == 'gestor':
                return redirect('gestor_page')
            elif user.user_type == 'diretor':
                return redirect('diretor_page')
            elif user.user_type == 'presidente':
                return redirect('presidente_page')
            elif user.user_type == 'rh':
                return redirect('rh_page')
            elif user.user_type == 'complice':
                return redirect('complice_page')
            else:
                return redirect('home')
    else:
        form = AuthenticationForm(request)
    return render(request, 'conexaorh/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def gestor_page(request):
    if request.user.user_type != 'gestor':
        return HttpResponseForbidden("Acesso negado! Apenas gestores podem acessar esta página.")

    form_rp = RequisicaoPessoalForm()
    form_movimentacao = MovimentacaoPessoalForm()
    form_rd = RequisicaoDesligamentoForm()
   
    movimentacao = MovimentacaoPessoal.objects.filter(
        ~Q(assinatura_complice__isnull=True),
        ~Q(assinatura_complice=""),
        gestor_proposto=request.user.username
    )
    form = GestorPropostoApprovalForm()

    if request.method == "POST":
        if "submit_rp" in request.POST:
            form_rp = RequisicaoPessoalForm(request.POST)
            if form_rp.is_valid():
                rp = form_rp.save(commit=False)
                rp.gestor = request.user
                rp.save()
                return redirect("gestor_page")

        elif "submit_movimentacao" in request.POST:
            form_movimentacao = MovimentacaoPessoalForm(request.POST)
            if form_movimentacao.is_valid():
                movimentacao = form_movimentacao.save(commit=False)
                movimentacao.save()
                return redirect("gestor_page")
        
        if "submit_aprovacao_mov" in request.POST:
            registro_id = request.POST.get("registro_id")
            if not registro_id:
                return HttpResponseBadRequest("ID do registro não informado.")
            
            registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)
            form = GestorPropostoApprovalForm(request.POST, request.FILES, instance=registro)

            if form.is_valid():
                assinatura = form.cleaned_data['assinatura_gestor_proposto']
                if assinatura:
                    registro.assinatura_gestor_proposto = assinatura
                    if registro.data_autorizacao_gestor_proposto is None:
                        registro.data_autorizacao_gestor_proposto = now()
                        registro.dias_para_autorizacao_gestor_proposto = (
                            now().date() - registro.data_solicitacao.date()
                        ).days
                    registro.save()
            return redirect("gestor_page")
            
        elif "submit_rd" in request.POST:
            form_rd = RequisicaoDesligamentoForm(request.POST)
            if form_rd.is_valid():
                rd = form_rd.save(commit=False)
                rd.save()
                return redirect("gestor_page")
            

    return render(request, "conexaorh/gestor.html", {
        "form_rp": form_rp,
        "form_movimentacao": form_movimentacao,
        "form_rd": form_rd,
        "movimentacao": movimentacao,
        "form": form,
        "usuario": request.user,
    })

@login_required
def complice_page(request):
    if request.user.user_type != 'complice':
        return HttpResponseForbidden("Acesso negado!")

    movimentacao = MovimentacaoPessoal.objects.all()

    form = CompliceApprovalForm()

    if request.method == 'POST':
        registro_id = request.POST.get('registro_id')
        if not registro_id:
            return HttpResponseBadRequest("ID do registro não informado.")
        
        registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)
        form = CompliceApprovalForm(request.POST, request.FILES, instance=registro)

        if form.is_valid():
            registro = form.save(commit=False)
            if registro.assinatura_complice and registro.data_autorizacao_complice is None:
                registro.data_autorizacao_complice = now()
                registro.dias_para_autorizacao_complice = (
                    registro.data_autorizacao_complice.date()
                    - registro.data_solicitacao.date()
                ).days
            registro.save()
            return redirect('complice_page')

    return render(request, 'conexaorh/complice.html', {
        'movimentacao': movimentacao,
        'usuario': request.user,
        'form': form
    })




@login_required
def diretor_page(request):
    if request.user.user_type != "diretor":
        return HttpResponseForbidden("Acesso negado! Apenas diretores podem acessar esta página.")

    rp = RequisicaoPessoal.objects.all()
    movimentacao = MovimentacaoPessoal.objects.filter(~Q(assinatura_gestor_proposto__isnull=True), ~Q(assinatura_gestor_proposto=""))
    rd = RequisicaoDesligamento.objects.all()


    form = DiretorForm()

    if request.method == "POST":
        registro_id = request.POST.get("registro_id")
        tipo_registro = request.POST.get("tipo_registro")  # Identifica o tipo

        if tipo_registro == "rp":
            registro = get_object_or_404(RequisicaoPessoal, id=registro_id)
        elif tipo_registro == "movimentacao":
            registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)
        elif tipo_registro == "rd":
            registro = get_object_or_404(RequisicaoDesligamento, id=registro_id)
        else:
            return HttpResponseBadRequest("Tipo de registro inválido.")

        form = DiretorForm(request.POST, request.FILES, instance=registro)

        if form.is_valid():
            registro = form.save(commit=False)
            # se acabou de assinar
            if registro.assinatura_diretor and registro.data_autorizacao_diretor is None:
                registro.data_autorizacao_diretor = now()
                registro.dias_para_autorizacao_diretor = (
                    registro.data_autorizacao_diretor.date()
                    - registro.data_solicitacao.date()
                ).days
            registro.save()
            return redirect("diretor_page")

    return render(request, "conexaorh/diretor.html", {"rp": rp, "movimentacao": movimentacao, "rd": rd, "usuario": request.user, "form": form})


@login_required
def presidente_page(request):
    if request.user.user_type != "presidente":
        return HttpResponseForbidden("Acesso negado! Apenas presidentes podem acessar esta página.")

    rp = RequisicaoPessoal.objects.filter(~Q(assinatura_diretor__isnull=True), ~Q(assinatura_diretor=""))
    movimentacao = MovimentacaoPessoal.objects.filter(~Q(assinatura_diretor__isnull=True), ~Q(assinatura_diretor=""))
    rd = RequisicaoDesligamento.objects.filter(~Q(assinatura_diretor__isnull=True), ~Q(assinatura_diretor=""))

    form = PresidenteForm()
    if request.method == "POST":
        registro_id = request.POST.get("registro_id")
        tipo_registro = request.POST.get("tipo_registro")

        if tipo_registro == "rp":
            registro = get_object_or_404(RequisicaoPessoal, id=registro_id)
        elif tipo_registro == "movimentacao":
            registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)
        elif tipo_registro == "rd":
            registro = get_object_or_404(RequisicaoDesligamento, id=registro_id)
        else:
            return HttpResponseBadRequest("Tipo de registro inválido.")

        form = PresidenteForm(request.POST, request.FILES, instance=registro)

        if form.is_valid():
            registro = form.save(commit=False)
            # se acabou de assinar
            if registro.assinatura_presidente and registro.data_autorizacao_presidente is None:
                registro.data_autorizacao_presidente = now()
                registro.dias_para_autorizacao_presidente = (
                    registro.data_autorizacao_presidente.date()
                    - registro.data_solicitacao.date()
                ).days
            registro.save()
            
            return redirect("presidente_page")

    return render(
        request, 
        "conexaorh/presidente.html", 
        {"rp": rp, "movimentacao": movimentacao, "rd": rd, "usuario": request.user, "form": form}
    )

@login_required
def rh_page(request):
    if request.user.user_type != "rh":
        return HttpResponseForbidden("Acesso negado! Apenas usuários do RH podem acessar esta página.")

    # Buscar registros dos dois modelos que já foram aprovados pelo presidente
    rp = RequisicaoPessoal.objects.filter(~Q(assinatura_presidente__isnull=True), ~Q(assinatura_presidente=""))
    movimentacao = MovimentacaoPessoal.objects.filter(~Q(assinatura_presidente__isnull=True), ~Q(assinatura_presidente=""))
    rd = RequisicaoDesligamento.objects.filter(~Q(assinatura_presidente__isnull=True), ~Q(assinatura_presidente=""))

    form = RHForm()

    if request.method == "POST":
        registro_id = request.POST.get("registro_id")
        tipo_registro = request.POST.get("tipo_registro")

        if tipo_registro == "rp":
            registro = get_object_or_404(RequisicaoPessoal, id=registro_id)
        elif tipo_registro == "movimentacao":
            registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)
        elif tipo_registro == "rd":
            registro = get_object_or_404(RequisicaoDesligamento, id=registro_id)
        else:
            return HttpResponseBadRequest("Tipo de registro inválido.")

        form = RHForm(request.POST, request.FILES, instance=registro)

        if form.is_valid():
            registro = form.save(commit=False)
            # se acabou de assinar
            if registro.assinatura_rh and registro.data_autorizacao_rh is None:
                registro.data_autorizacao_rh = now()
                registro.dias_para_autorizacao_rh = (
                    registro.data_autorizacao_rh.date()
                    - registro.data_solicitacao.date()
                ).days
            registro.save()

            return redirect("rh_page")

    return render(
        request,
        "conexaorh/rh.html",
        {"rp": rp, "movimentacao": movimentacao, "rd": rd, "usuario": request.user, "form": form}
    )

def preencher_excel_rp(rp):
    wb = load_workbook("conexaorh/static/rp_modelo.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=60):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    template_processo_seletivo = "[  ] INTERNO [  ] MISTO [  ] EXTERNO [  ] SIGILOSO"

    processo_seletivo = [(v or "").strip().casefold() for v in (rp.processo_seletivo or "").split(",")]

    for opcao_processo_seletivo in ["INTERNO", "INTERNO", "EXTERNO", "SIGILOSO"]:
        if opcao_processo_seletivo.casefold() in processo_seletivo:
            template_processo_seletivo = template_processo_seletivo.replace(f"[  ] {opcao_processo_seletivo}", f"[X] {opcao_processo_seletivo}")

    # Texto base da célula no Excel
    template_sexo = "[  ] Masculino   [  ] Feminino   [  ] Indiferente"

    # Pega os valores separados por vírgula, normaliza com strip() e title()
    sexos = [(s or "").strip().title() for s in (rp.sexo or "").split(",")]

    # Para cada opção, substitui se estiver presente na lista
    for opcao in ["Masculino", "Feminino", "Indiferente"]:
        if opcao in sexos:
            template_sexo = template_sexo.replace(f"[  ] {opcao}", f"[X] {opcao}")
    
    template_beneficios = "[  ] Plano Unimed Cnu [  ] Plano Saúde Hapvida [  ] Alelo Mobilidade [  ] Seguro De Vida"

    beneficios = [(b or "").strip().title() for b in (rp.beneficios or "").split(",")]

    for opcao_beneficios in ["Plano Unimed Cnu", "Plano Saúde Hapvida", "Alelo Mobilidade", "Seguro De Vida"]:
        for opcao_beneficios in beneficios:
            template_beneficios = template_beneficios.replace(f"[  ] {opcao_beneficios}", f"[X] {opcao_beneficios}") 
    
    template_viagem = "[  ] Sim [  ] Não"

    viagens = [(v or "").strip().casefold() for v in (rp.exige_viagem or "").split(",")]

    for opcao in ["Sim", "Não"]:
        if opcao.casefold() in viagens:
            template_viagem = template_viagem.replace(f"[  ] {opcao}", f"[X] {opcao}")

    template_cnh = "[  ] Sim [  ] Não"

    cnh = [(v or "").strip().casefold() for v in (rp.cnh or "").split(",")]

    for opcao_cnh in ["Sim", "Não"]:
        if opcao_cnh.casefold() in cnh:
            template_cnh = template_cnh.replace(f"[  ] {opcao_cnh}", f"[X] {opcao_cnh}")

    ws["I4"] = rp.data_solicitacao.strftime("%d/%m/%Y") if rp.data_solicitacao else ""
    ws["G7"] = rp.requisitante or ""
    ws["X7"] = rp.cargo or ""
    ws["G9"] = rp.salario or ""
    ws["U9"] = rp.adicionais or ""
    ws["AF9"] = rp.quantidade_vagas or ""
    ws["G11"] = rp.horario_trabalho_inicio.strftime("%H:%M") if rp.horario_trabalho_inicio else ""
    ws["P11"] = rp.tipo_ponto or ""
    ws["X11"] = rp.inicio_contrato.strftime("%d/%m/%Y") if rp.inicio_contrato else ""
    ws["AF11"] = rp.termino_contrato.strftime("%d/%m/%Y") if rp.termino_contrato else ""
    ws["G13"] = rp.tipo_contratacao or ""
    ws["H16"] = template_beneficios
    ws["U13"] = rp.motivo_contracao or ""
    ws["H17"] = rp.subtituicao or ""
    ws["AE17"] = rp.matricula or "" 
    ws["I19"] = rp.justificativa_substituicao or ""
    ws["AC19"] = rp.justificativa_outros or ""
    ws["E22"] = template_processo_seletivo
    ws["Y22"] = rp.localidade or ""
    ws["AG22"] = rp.base or ""
    ws["O24"] = template_sexo
    ws["AD24"] = template_viagem
    ws["C26"] = template_cnh
    ws["K26"] = rp.tipo_cnh or ""
    ws["S26"] = rp.outros_cnh or ""
    ws["G28"] = rp.departamento or ""
    ws["Y28"] = rp.escolaridade or ""
    ws["G30"] = rp.gestor_imediato or ""
    ws["Y30"] = rp.centro_custo or ""
    ws["G32"] = rp.cursos or ""
    ws["G34"] = rp.experiencias or ""
    ws["A38"] = rp.habilidades_comportamentais or ""
    ws["S38"] = rp.principais_atribuicoes or ""
    ws["G53"] = rp.candidato_aprovado or ""
    ws["AB53"] = rp.n_rp or ""

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, rp.assinatura_diretor, "S61")
    adicionar_imagem_excel(ws, rp.assinatura_presidente, "AB61")
    adicionar_imagem_excel(ws, rp.assinatura_rh, "A61")

    # Salva em memória
    buffer = io.BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_rp_excel(request, registro_id):
    registro = get_object_or_404(RequisicaoPessoal, id=registro_id)

    excel_bytes = preencher_excel_rp(registro)
    nome_arquivo = f"RP_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response

def preencher_excel_mov(mov):
    wb = load_workbook("conexaorh/static/mov_modelo.xlsx")
    ws = wb.active

    def adicionar_imagem_excel(ws, imagem_field, celula, largura=120, altura=60):
        if imagem_field and imagem_field.name:
            try:
                img_pillow = Image.open(imagem_field)
                img_buffer = io.BytesIO()
                img_pillow.save(img_buffer, format="PNG")
                img_buffer.seek(0)

                img_excel = XLImage(img_buffer)
                img_excel.width = largura
                img_excel.height = altura
                ws.add_image(img_excel, celula)
            except Exception as e:
                print(f"Erro ao adicionar imagem em {celula}: {e}")

    try:
        usuario = User.objects.get(username=mov.gestor_proposto)  # Se armazenar username
        nome_gestor = usuario.get_full_name() or usuario.first_name or usuario.username
    except User.DoesNotExist:
        nome_gestor = mov.gestor_proposto  # Usa o texto mesmo se não encontrar

    template_tipo_movimentacao = (
        "[  ] PROMOÇÃO HORIZONTAL   [  ] PROMOÇÃO VERTICAL   [  ] TRANSFERÊNCIA DE ÁREA\n"
        "[  ] TRANSFERÊNCIA DE LOCALIDADE   [  ] TRANSFERÊNCIA DE CONTRATO   [  ] ENQUADRAMENTO")

    tipo_movimentacao = [(v or "").strip().casefold() for v in (mov.tipo_movimentacao or "").split(",")]

    for opcao_tipo_movimentacao in ["PROMOÇÃO HORIZONTAL", "PROMOÇÃO VERTICAL", "TRANSFERÊNCIA DE ÁREA", "TRANSFERÊNCIA DE LOCALIDADE", "TRANSFERÊNCIA DE CONTRATO", "ENQUADRAMENTO"]:
        if opcao_tipo_movimentacao.casefold() in tipo_movimentacao:
            template_tipo_movimentacao = template_tipo_movimentacao.replace(f"[  ] {opcao_tipo_movimentacao}", f"[X] {opcao_tipo_movimentacao}")

    template_tipo_adicional = (
        "[  ] PERICULOSIDADE [  ] INSALUBRIDADE [  ] AJUDA DE CUSTO [  ] ADICIONAL"
    )

    tipo_adicional = [(v or "").strip().casefold() for v in (mov.tipo_adicional or "").split(",")]

    for opcao_tipo_adicional in ["PERICULOSIDADE", "INSALUBRIDADE", "AJUDA DE CUSTO", "ADICIONAL"]:
        if opcao_tipo_adicional.casefold() in tipo_adicional:
            template_tipo_adicional = template_tipo_adicional.replace(f"[  ] {opcao_tipo_adicional}", f"[X] {opcao_tipo_adicional}")

    template_jutificativa_movimentacao = (
        "[  ] REESTRUTURAÇÃO DO DEPARTAMENTO (EMPRESA/UNIDADE) [  ] OPORTUNIDADE DE ASCENÇÃO\n"
        "[  ] INADEQUAÇÃO À ATIVIDADE DO DEPARTAMENTO "
    )

    jutificativa_movimentacao = [(v or "").strip().casefold() for v in (mov.jutificativa_movimentacao or "").split(",")]

    for opcao_jutificativa_movimentacao in ["REESTRUTURAÇÃO DO DEPARTAMENTO (EMPRESA/UNIDADE)", "INADEQUAÇÃO À ATIVIDADE DO DEPARTAMENTO", "OPORTUNIDADE DE ASCENÇÃO"]:
        if opcao_jutificativa_movimentacao.casefold() in jutificativa_movimentacao:
            template_jutificativa_movimentacao = template_jutificativa_movimentacao.replace(f"[  ] {opcao_jutificativa_movimentacao}", f"[X] {opcao_jutificativa_movimentacao}")

    template_substituicao = (
        "[  ] SIM [  ] NÃO  "
    )

    substituicao = [(v or "").strip().casefold() for v in (mov.substituicao or "").split(",")]

    for opcao_substituicao in ["SIM", "NÃO"]:
        if opcao_substituicao.casefold() in substituicao:
            template_substituicao = template_substituicao.replace(f"[  ] {opcao_substituicao}", f"[X] {opcao_substituicao}")
      

    ws["C5"] = mov.numero or ""
    ws["O5"] = mov.data_solicitacao.strftime("%d/%m/%Y") if mov.data_solicitacao else ""
    ws["AC5"] = mov.unidade or ""
    ws["A9"] = template_tipo_movimentacao
    ws["A9"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["A15"] = "OUTRO (CITAR)"
    ws["A15"].font = Font(bold=True)
    ws["H15"] = mov.outro_tipo or ""
    ws["H17"] = mov.colaborador_movimentado or ""
    ws["F19"] = mov.matricula or ""
    ws["S19"] = mov.data_admissao.strftime("%d/%m/%Y") if mov.data_admissao else ""
    ws["AD19"] = mov.outro_info or ""
    ws["A22"] = mov.localidade_atual or ""
    ws["J22"] = mov.cargo_atual or ""
    ws["S22"] = mov.departamento_atual or ""
    ws["AB22"] = mov.salario_atual or ""
    ws["F24"] = mov.gestor_atual or ""
    ws["AD24"] = mov.centro_custo_atual or ""
    ws["A27"] = mov.localidade_proposta or ""
    ws["J27"] = mov.cargo_proposto or ""
    ws["S27"] = mov.departamento_proposto or ""
    ws["AB27"] = mov.salario_proposto or ""
    ws["F29"] = nome_gestor
    ws["AD29"] = mov.centro_custo_proposto or ""
    ws["S32"] = mov.data_movimentacao.strftime("%d/%m/%Y") if mov.data_movimentacao else ""
    ws["A34"] = template_tipo_adicional
    ws["F36"] = mov.tipo_ajuda_custo or ""
    ws["R36"] = mov.valor_ajuda or ""
    ws["AA36"] = mov.periodo or ""
    ws["A39"] = template_jutificativa_movimentacao
    ws["A39"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws["H43"] = mov.outro_justificativa or ""
    ws["A43"] = "OUTRO (CITAR)"
    ws["A43"].font = Font(bold=True)
    ws["U45"] = template_substituicao
    ws["H46"] = mov.comentarios or ""
    ws["A32"] = mov.data_solicitacao.strftime("%d/%m/%Y") if mov.data_solicitacao else ""
    ws["A60"] = mov.data_autorizacao_complice.strftime("%d/%m/%Y") if mov.data_autorizacao_complice else ""
    ws["J65"] = mov.data_autorizacao_gestor_proposto.strftime("%d/%m/%Y") if mov.data_autorizacao_gestor_proposto else ""
    ws["S65"] = mov.data_autorizacao_diretor.strftime("%d/%m/%Y") if mov.data_autorizacao_diretor else ""
    ws["AB65"] = mov.data_autorizacao_presidente.strftime("%d/%m/%Y") if mov.data_autorizacao_presidente else ""
    ws["S60"] = mov.data_autorizacao_rh.strftime("%d/%m/%Y") if mov.data_autorizacao_rh else ""
    

    # Inserção das assinaturas
    adicionar_imagem_excel(ws, mov.assinatura_complice, "A56")
    adicionar_imagem_excel(ws, mov.assinatura_gestor_proposto, "J68")
    adicionar_imagem_excel(ws, mov.assinatura_diretor, "S68")
    adicionar_imagem_excel(ws, mov.assinatura_presidente, "AB68")
    adicionar_imagem_excel(ws, mov.assinatura_rh, "S56")


    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()

@login_required
def download_mov_excel(request, registro_id):
    registro = get_object_or_404(MovimentacaoPessoal, id=registro_id)

    excel_bytes = preencher_excel_mov(registro)
    nome_arquivo = f"Movimentação_{registro.id}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response
     
     
     
       
     
     
     

     
     
     
     
     
     
     
      
     
     
     
     
     
     
    